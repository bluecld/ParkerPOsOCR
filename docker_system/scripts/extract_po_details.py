"""
Extract detailed information from PO PDF file
"""

import fitz
import pytesseract
from PIL import Image
import io
import re
import json
import os
from pathlib import Path

# Use system tesseract in container (no hardcoded Windows path)

# Global cache for part numbers validation
_PART_NUMBERS_CACHE = None
_PART_TO_FULL_CACHE = None

def load_part_numbers_for_validation():
    """Load part numbers from Excel file for validation (with caching)"""
    global _PART_NUMBERS_CACHE, _PART_TO_FULL_CACHE
    
    if _PART_NUMBERS_CACHE is not None:
        return _PART_NUMBERS_CACHE, _PART_TO_FULL_CACHE
        
    try:
        # Try to import openpyxl for Excel reading
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("Info: openpyxl not available, part number validation disabled")
            _PART_NUMBERS_CACHE = set()
            _PART_TO_FULL_CACHE = {}
            return _PART_NUMBERS_CACHE, _PART_TO_FULL_CACHE
            
        excel_path = "/app/docs/PartNumbers.xlsx"
        
        if not os.path.exists(excel_path):
            print(f"Warning: Part numbers file not found at {excel_path}")
            _PART_NUMBERS_CACHE = set()
            _PART_TO_FULL_CACHE = {}
            return _PART_NUMBERS_CACHE, _PART_TO_FULL_CACHE
            
        # Load Excel file with openpyxl
        workbook = load_workbook(excel_path, read_only=True)
        worksheet = workbook.active
        
        part_numbers = set()
        part_to_full = {}
        
        # Assume first row is headers, start from row 2
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # Check if first column (Part Number) has data
                full_part = str(row[0]).strip()
                if '*' in full_part:
                    clean_part = full_part.split('*')[0].upper().strip()
                else:
                    clean_part = full_part.upper().strip()
                
                part_numbers.add(clean_part)
                part_to_full[clean_part] = full_part
        
        workbook.close()
        _PART_NUMBERS_CACHE = part_numbers
        _PART_TO_FULL_CACHE = part_to_full
        print(f"Loaded {len(part_numbers)} part numbers for validation from Excel")
        
    except Exception as e:
        print(f"Warning: Could not load part numbers for validation: {e}")
        _PART_NUMBERS_CACHE = set()
        _PART_TO_FULL_CACHE = {}
        
    return _PART_NUMBERS_CACHE, _PART_TO_FULL_CACHE

def fix_common_ocr_errors(text):
    """Fix common OCR character recognition errors in part numbers"""
    if not text:
        return [text]
        
    # Common OCR mistakes for part numbers
    fixes = {
        'Q': '9',  # Q often misread as 9 (WAQ04 -> WA904)
        'O': '0',  # O often misread as 0 (97O000 -> 970000)  
        'I': '1',  # I often misread as 1
        'S': '5',  # S sometimes misread as 5
        'Z': '2',  # Z sometimes misread as 2
        'G': '6',  # G sometimes misread as 6
        'B': '8',  # B sometimes misread as 8
    }
    
    corrected_candidates = [text]
    
    # Generate candidates by fixing one character at a time
    for i, char in enumerate(text.upper()):
        if char in fixes:
            candidate = text[:i] + fixes[char] + text[i+1:]
            corrected_candidates.append(candidate)
            
        # Also try the reverse (number to letter) for edge cases
        for fix_char, fix_replacement in fixes.items():
            if char == fix_replacement:
                reverse_candidate = text[:i] + fix_char + text[i+1:]
                corrected_candidates.append(reverse_candidate)
    
    return corrected_candidates

def validate_part_number_with_reference(ocr_part_number):
    """
    Validate and correct part number using reference database
    PRESERVES the original OP code from OCR - only corrects the part number portion
    Returns: (corrected_part_number, confidence_score, correction_info)
    """
    if not ocr_part_number:
        return ocr_part_number, 0.0, "Empty input"
        
    part_numbers, part_to_full = load_part_numbers_for_validation()
    
    if not part_numbers:
        # No reference data available, return as-is
        return ocr_part_number, 0.5, "No reference data available"
        
    # Clean the OCR text and preserve the original OP code
    clean_ocr = ocr_part_number.strip().upper()
    if '*' in clean_ocr:
        clean_part = clean_ocr.split('*')[0]
        original_op_part = '*' + clean_ocr.split('*')[1]  # PRESERVE original OP
    else:
        clean_part = clean_ocr
        original_op_part = ''
        
    # 1. Exact match check
    if clean_part in part_numbers:
        # Use original OP code, not database OP code
        corrected_full = clean_part + original_op_part
        return corrected_full, 1.0, f"Exact match: {clean_part}"
        
    # 2. Try OCR error corrections
    candidates = fix_common_ocr_errors(clean_part)
    for candidate in candidates:
        if candidate.upper() in part_numbers:
            # Use original OP code, not database OP code
            corrected_full = candidate.upper() + original_op_part
            return corrected_full, 0.95, f"OCR correction: {clean_part} → {candidate.upper()}"
            
    # 3. Fuzzy matching for close matches
    from difflib import get_close_matches
    try:
        close_matches = get_close_matches(clean_part, part_numbers, n=3, cutoff=0.85)
        if close_matches:
            best_match = close_matches[0]
            # Use original OP code, not database OP code
            corrected_full = best_match + original_op_part
            return corrected_full, 0.8, f"Fuzzy match: {clean_part} → {best_match}"
    except:
        pass
        
    # 4. Return original if no match found
    return ocr_part_number, 0.3, f"No match found in reference database"

def extract_text_from_pdf(pdf_path):
    """Extract all text from PDF using OCR with better accuracy - processes ALL pages in the PDF"""
    doc = fitz.open(pdf_path)
    all_text = ""
    
    # Process EVERY page in the PDF (this already scanned all pages)
    for page_num in range(len(doc)):
        page = doc[page_num]
        
        # Try to extract text directly first
        page_text = page.get_text()
        
        # If no text found or minimal text, use OCR
        if len(page_text.strip()) < 50:
            # Use adaptive resolution based on source quality
            # For 300dpi scans, use higher scaling; for 600dpi, moderate scaling
            matrices = [
                fitz.Matrix(4, 4),  # Higher res for 300dpi scans
                fitz.Matrix(3, 3),  # Standard high res
                fitz.Matrix(2, 2)   # Lower res for already high-quality scans
            ]
            
            best_text = ""
            best_length = 0
            
            for matrix in matrices:
                try:
                    pix = page.get_pixmap(matrix=matrix)
                    img_data = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_data)).convert('L')
                    
                    # Enhanced OCR configs for different scan qualities
                    configs = [
                        r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz/-*().,: ',
                        r'--oem 3 --psm 4',
                        r'--oem 3 --psm 3',
                        r'--oem 1 --psm 6',  # Legacy engine for difficult scans
                        r'--oem 2 --psm 6'   # Cube engine
                    ]
                    
                    for config in configs:
                        try:
                            ocr_text = pytesseract.image_to_string(img, config=config)
                            if len(ocr_text) > best_length:
                                best_text = ocr_text
                                best_length = len(ocr_text)
                        except:
                            continue
                    
                    # If we got decent results, don't try more matrices
                    if best_length > 200:
                        break
                        
                except Exception as e:
                    print(f"OCR matrix {matrix} failed: {e}")
                    continue
            
            page_text = best_text if best_text else page_text
        
        all_text += f"PAGE {page_num + 1}:\n{page_text}\n\n"
    
    doc.close()
    return all_text

def extract_production_order(text):
    """Extract production order number - flexible pattern matching from comprehensive text"""
    # Multiple patterns to try, in order of preference
    patterns = [
        r'12\d{7}',           # Parker format: 12 followed by 7 digits (125157207)
        r'\b1[0-9]{8}\b',     # 9-digit numbers starting with 1
        r'\b[1-9]\d{7,9}\b',  # 8-10 digit numbers starting with non-zero
        r'Production\s+Order[:\s]+(\d{8,10})',  # "Production Order: NNNNNN"
        r'PO[:\s]+(\d{8,10})', # "PO: NNNNNN"
        r'WO[:\s]+(\d{8,10})'  # "WO: NNNNNN" (Work Order)
    ]
    
    # Search through all sections of the comprehensive text
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # Return the first match, handling both direct matches and group captures
            result = matches[0]
            # Validate the number looks reasonable (not too many repeated digits)
            if isinstance(result, str) and len(result) >= 8:
                # Check for obviously invalid patterns (all same digit, etc.)
                if not re.match(r'^(\d)\1+$', result):  # Not all same digit
                    return result
    
    return None

def extract_revision(text):
    """Extract revision number following 'REV' - clean to 1-3 characters"""
    # Prefer a single letter/letter+digit right after REV to avoid grabbing unrelated tokens
    pattern = r'\bREV\s*([A-Z](?:[0-9])?)\b'
    matches = re.findall(pattern, text, re.IGNORECASE)
    if matches:
        revision = matches[0].upper()
        # Accept 1-2 char revisions like C, D1; otherwise trim to 1 char
        if len(revision) in (1, 2):
            return revision
        return revision[:1]
    return None

def extract_part_number(text, production_order):
    """Extract part number from comprehensive text (PO pages + first router page)"""
    if not production_order:
        return None
    
    lines = text.split('\n')
    
    # Look for production order and then search nearby lines for part number
    # Enhanced to work with comprehensive text from multiple sections
    for i, line in enumerate(lines):
        if production_order in line:
            # Search in a wider context around the production order
            start_idx = max(0, i-15)  # Increased search range for comprehensive text
            end_idx = min(len(lines), i+10)
            context_lines = lines[start_idx:end_idx]
            
            # Pattern 1: Support both digit-digit (157710-30) and alpha+digits with optional dash (WA904-8)
            for j, context_line in enumerate(context_lines):
                # Look for part with OP on same line first
                dash_pattern_match = re.search(r'(?:\b(\d+[-_]\d+)\b|\b([A-Z]{1,4}\d{2,6}(?:[-_]\d{1,3})?)\b)\s*[*]?([Oo]p\d+)', context_line, re.IGNORECASE)
                if dash_pattern_match:
                    part_base = dash_pattern_match.group(1) or dash_pattern_match.group(2)
                    part_base = part_base.upper()
                    op_part = dash_pattern_match.group(3).upper()
                    raw_part = f"{part_base}*{op_part}"
                    
                    # Use database validation first (includes OCR corrections)
                    validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                    if confidence > 0.8:  # High confidence correction from database
                        print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                        return validated_part
                    
                    # Fallback to simple OCR corrections if database validation failed
                    corrected_candidates = fix_common_ocr_errors(part_base)
                    if corrected_candidates and len(corrected_candidates) > 1:
                        # Use the first corrected candidate (after the original)
                        corrected_part = f"{corrected_candidates[1]}*{op_part}"
                        print(f"OCR correction applied: {raw_part} → {corrected_part}")
                        return corrected_part
                    
                    return raw_part
                
                # Look for a part pattern that might have OP on next line or nearby
                dash_match = re.search(r'(\d+[-_]\d+|[A-Z]{1,4}\d{2,6}(?:[-_]\d{1,3})?)', context_line, re.IGNORECASE)
                if dash_match:
                    part_base = dash_match.group(1).upper()
                    
                    # Search in multiple subsequent lines for OP pattern
                    for k in range(j+1, min(j+5, len(context_lines))):
                        if k < len(context_lines):
                            search_line = context_lines[k]
                            op_match = re.search(r'[*]?([Oo]p\d+)', search_line, re.IGNORECASE)
                            if op_match:
                                op_part = op_match.group(1).upper()
                                raw_part = f"{part_base}*{op_part}"
                                
                                # Use database validation first (includes OCR corrections)
                                validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                                if confidence > 0.8:  # High confidence correction from database
                                    print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                                    return validated_part
                                
                                # Fallback to simple OCR corrections if database validation failed
                                corrected_candidates = fix_common_ocr_errors(part_base)
                                if corrected_candidates and len(corrected_candidates) > 1:
                                    # Use the first corrected candidate (after the original)
                                    corrected_part = f"{corrected_candidates[1]}*{op_part}"
                                    print(f"OCR correction applied: {raw_part} → {corrected_part}")
                                    return corrected_part
                                
                                return raw_part
                    
                    # If no OP found, also search in the broader context around this part number
                    context_text = ' '.join(context_lines[max(0, j-3):min(len(context_lines), j+8)])
                    op_match = re.search(rf'{re.escape(part_base)}.*?([Oo]p\d+)', context_text, re.IGNORECASE)
                    if op_match:
                        op_part = op_match.group(1).upper()
                        raw_part = f"{part_base}*{op_part}"
                        
                        # Use database validation first (includes OCR corrections)
                        validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                        if confidence > 0.8:  # High confidence correction from database
                            print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                            return validated_part
                        
                        # Fallback to simple OCR corrections if database validation failed
                        corrected_candidates = fix_common_ocr_errors(part_base)
                        if corrected_candidates and len(corrected_candidates) > 1:
                            # Use the first corrected candidate (after the original)
                            corrected_part = f"{corrected_candidates[1]}*{op_part}"
                            print(f"OCR correction applied: {raw_part} → {corrected_part}")
                            return corrected_part
                        
                        return raw_part
                    
                    # Store this as a candidate in case we don't find anything better
                    candidate_part = part_base
            
            # Pattern 2: Look for just digits (like 521350) near production order
            for j, context_line in enumerate(context_lines):
                # Look for 6-digit numbers that appear before Op or near production order
                digit_matches = re.findall(r'\b(\d{6})\b', context_line)
                for digit_match in digit_matches:
                    # Check if this appears near an Op pattern
                    context_text = ' '.join(context_lines)
                    if re.search(rf'{digit_match}.*?[Oo]p\d+', context_text, re.IGNORECASE):
                        # Find the Op number
                        op_match = re.search(rf'{digit_match}.*?([Oo]p\d+)', context_text, re.IGNORECASE)
                        if op_match:
                            op_part = op_match.group(1).upper()
                            raw_part = f"{digit_match}*{op_part}"
                            
                            # Validate and correct the part number
                            validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                            if confidence > 0.8:  # High confidence correction
                                print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                                return validated_part
                            else:
                                return raw_part
                        else:
                            # Validate standalone digit part
                            validated_part, confidence, correction_info = validate_part_number_with_reference(digit_match)
                            if confidence > 0.8:
                                print(f"Part number validation: {digit_match} → {validated_part} ({correction_info})")
                                return validated_part
                            else:
                                return digit_match
            
            # Pattern 3: Look for part patterns without OP, but search harder for OP
            candidate_part = None
            for j, context_line in enumerate(context_lines):
                dash_matches = re.findall(r'(\d+[-_]\d+|[A-Z]{1,4}\d{2,6}(?:[-_]\d{1,3})?)', context_line, re.IGNORECASE)
                if dash_matches:
                    candidate_part = dash_matches[-1].upper()  # Take the last/closest one to production order
                    
                    # Search the entire context for an OP that might go with this part
                    full_context = ' '.join(context_lines)
                    # Look for OP20, OP30, etc. anywhere in the context
                    op_matches = re.findall(r'[Oo]p\d+', full_context, re.IGNORECASE)
                    if op_matches:
                        # Take the first OP found (most likely to be associated)
                        op_part = op_matches[0].upper()
                        raw_part = f"{candidate_part}*{op_part}"
                        
                        # Validate and correct the part number
                        validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                        if confidence > 0.8:  # High confidence correction
                            print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                            return validated_part
                        else:
                            return raw_part
                    else:
                        # Return the part without OP for now, but keep looking
                        break
            
            # If we found a candidate part but no OP, add default *OP20
            if candidate_part:
                raw_part = f"{candidate_part}*OP20"
                # Validate and correct the part number using reference database
                validated_part, confidence, correction_info = validate_part_number_with_reference(raw_part)
                if confidence > 0.8:  # High confidence correction
                    print(f"Part number validation: {raw_part} → {validated_part} ({correction_info})")
                    return validated_part
                else:
                    return raw_part
            
            # Pattern 4: Look for standalone numbers near production order
            for j, context_line in enumerate(context_lines):
                if j < len(context_lines) - 2:  # Not the production order line itself
                    standalone_matches = re.findall(r'\b(\d{5,7})\b', context_line)
                    for match in standalone_matches:
                        # Skip obvious non-part numbers (production orders start with 12)
                        if not match.startswith('12') and not match.startswith('455'):
                            # Validate standalone part numbers too
                            validated_part, confidence, correction_info = validate_part_number_with_reference(match)
                            if confidence > 0.8:
                                print(f"Part number validation: {match} → {validated_part} ({correction_info})")
                                return validated_part
                            else:
                                return match
            
            break  # Found production order, stop looking
    
    # Apply validation to any discovered part number before returning
    if candidate_part:
        validated_part, confidence, correction_info = validate_part_number_with_reference(candidate_part)
        if confidence > 0.8:
            print(f"Part number validation: {candidate_part} → {validated_part} ({correction_info})")
            return validated_part
            
    return None

def extract_quantity_and_dock_date(text):
    """Extract quantity as whole number and dock date from the same context"""
    quantity = None
    dock_date = None
    
    lines = text.split('\n')
    
    # Strategy 0: Item-number anchored (prefer line item 10, else 20)
    def is_item_line(line: str, item_no: str) -> bool:
        # Consider it an item row if the item number appears near the left within first ~20 chars
        prefix = line[:20]
        return re.search(rf"\b{item_no}\b", prefix) is not None

    def extract_qty_from_line(line: str, next_lines: list[str] | None = None):
        nonlocal dock_date
        # Capture dock date from this line if present
        dm = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', line)
        if dm and not dock_date:
            dock_date = dm.group(1)

        # Find unit position if any
        unit_match = re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE)
        unit_pos = unit_match.start() if unit_match else None

        # Define a left region to search for quantity: before 'NET' or before first $price, or before unit
        left_region = line
        m_net = re.search(r'\bNET\b', line, re.IGNORECASE)
        if m_net:
            left_region = line[:m_net.start()]
        m_dollar = re.search(r'\$\s*\d', left_region)
        if m_dollar:
            left_region = left_region[:m_dollar.start()]
        if unit_pos is not None and unit_pos < len(left_region):
            left_region = left_region[:unit_pos]

        # Prefer the LARGEST .00 before the unit (quantity usually larger than unit price)
        mdec = [int(m.group(1)) for m in re.finditer(r'\b(\d{1,4})\.00\b', left_region)]
        if mdec:
            candidates = [v for v in mdec if 1 <= v <= 2000]
            if candidates:
                return max(candidates)

        # As a fallback, consider integers in left region but ignore leading item numbers like 10/20 at start
        int_positions = [(int(m.group(1)), m.start()) for m in re.finditer(r'\b(\d{1,4})\b', left_region)]
        filtered = []
        for val, pos in int_positions:
            # Skip common item numbers at the very left (e.g., '10', '20')
            if pos <= 3 and val in (10, 20, 30, 40):
                continue
            if 1 <= val <= 2000:
                filtered.append((pos, val))
        if filtered:
            # Choose the integer closest to the unit position (tends to be the quantity)
            filtered.sort(key=lambda t: t[0], reverse=True)
            return filtered[0][1]

        # If not found before the unit/NET, search near the unit (to the right as well)
        # Prefer quantities that end with .00 and are closest to the unit token
        whole_line = line
        candidates = []
        for m in re.finditer(r'\b(\d{1,4})\.00\b', whole_line):
            val = int(m.group(1))
            if 1 <= val <= 2000:
                dist = abs(m.start() - (unit_pos or m.start()))
                candidates.append((dist, m.start(), val))
        if candidates:
            # Choose the closest .00 to the unit, but if distances tie, take the larger value
            candidates.sort(key=lambda t: (t[0], -t[2], -t[1]))
            return candidates[0][2]

        # Vertical block fallback around this row (look into following few lines for Quantity/date/unit)
        if next_lines:
            window = [line] + next_lines[:5]
            window_text = "\n".join(window)
            # Try to find a date followed by a .00 and a unit token within a short window
            dm2 = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', window_text)
            if dm2 and not dock_date:
                dock_date = dm2.group(1)
            m2 = re.findall(r'\b(\d{1,3})\.00\b', window_text)
            if m2:
                # Choose the smallest plausible quantity (prices usually not .00)
                vals = [int(x) for x in m2 if 1 <= int(x) <= 500]
                if vals:
                    return min(vals)
        return None

    # Try line item 10 first
    for i, line in enumerate(lines):
        if is_item_line(line, '10'):
            q = extract_qty_from_line(line, lines[i+1:i+6])
            if q:
                quantity = q
                break
    # Fallback to line item 20
    if not quantity:
        for i, line in enumerate(lines):
            if is_item_line(line, '20'):
                q = extract_qty_from_line(line, lines[i+1:i+6])
                if q:
                    quantity = q
                    break
    
    # Strategy 1: Row-oriented extraction where a detail line contains qty(.00), unit, and date
    for i, line in enumerate(lines):
        # Must contain a unit and a date to be considered an item row
        if re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE) and re.search(r'\b\d{1,2}/\d{1,2}/\d{4}\b', line):
            # Find .00 or integer tokens near the unit token
            tokens = re.findall(r"[A-Za-z]+|\d+\.\d{2}|\d{1,4}/\d{1,2}/\d{4}|\d+", line)
            # Locate unit index
            unit_idx = next((idx for idx, t in enumerate(tokens) if re.fullmatch(r'(EA|LBS|PCS|EACH|PIECES?)', t, re.IGNORECASE)), None)
            # Extract date straight from the row
            date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', line)
            if date_match:
                dock_date = date_match.group(1)
            if unit_idx is not None:
                # Search up to 3 tokens before unit for a quantity number
                for pos in range(max(0, unit_idx-3), unit_idx):
                    t = tokens[pos]
                    m1 = re.fullmatch(r'(\d+)\.00', t)
                    m2 = re.fullmatch(r'(\d+)', t)
                    cand = None
                    if m1:
                        cand = int(m1.group(1))
                    elif m2:
                        cand = int(m2.group(1))
                    if cand is not None and 1 <= cand <= 1000:
                        quantity = cand
                        break
            if quantity:
                break

    # Strategy 1b: If the tokenization above missed, look in a small window around unit/date
    if not quantity:
        for i, line in enumerate(lines):
            if re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE):
                window = ' '.join(lines[max(0, i-1):i+2])
                # Prefer a small integer with .00 in the window
                candidates = [int(x) for x in re.findall(r'\b(\d{1,3})\.00\b', window)]
                if candidates:
                    quantity = min(candidates)
                    # Grab date nearby
                    dm = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', window)
                    if dm:
                        dock_date = dm.group(1)
                    break

    # Strategy 2: Smart context-aware extraction (respect column order; avoid Net Per)
    for i, line in enumerate(lines):
        if re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE) and re.search(r'\d', line):
            # Column anchors
            m_unit = re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE)
            pos_unit = m_unit.start() if m_unit else None
            pos_net = line.lower().find('net')
            # Prefer region before 'net' if it appears before the unit
            if pos_net != -1 and (pos_unit is None or pos_net < pos_unit):
                qty_region = line[:pos_net]
            elif pos_unit is not None:
                qty_region = line[:pos_unit]
            else:
                qty_region = line

            # Prefer rightmost .00 in the region as quantity
            decs = list(re.finditer(r'(\d+)\.00\b', qty_region))
            if decs:
                v = int(decs[-1].group(1))
                if 1 <= v <= 2000:
                    quantity = v
            else:
                ints = list(re.finditer(r'\b(\d{1,4})\b', qty_region))
                if ints:
                    v = int(ints[-1].group(1))
                    if 1 <= v <= 2000:
                        quantity = v

            dm = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', line)
            if dm and not dock_date:
                dock_date = dm.group(1)
            if quantity:
                break
    
    # Strategy 3: Vertical table format - detect around either the unit line or the Quantity label
    # Look for pattern: Quantity label with a nearby decimal and unit, or standalone unit line with nearby Quantity
    if not quantity:
        for i, line in enumerate(lines):
            # Case A: Unit on its own line, find Quantity label nearby
            if re.match(r'^\s*(EA|LBS|PCS|EACH|PIECES?)\s*$', line, re.IGNORECASE):
                area = lines[max(0, i-6):i+7]
                block = "\n".join(area)
                # Prefer .00 values in the block
                m = re.findall(r'\b(\d{1,3})\.00\b', block)
                if m:
                    vals = [int(x) for x in m if 1 <= int(x) <= 1000]
                    if vals:
                        quantity = min(vals)
                        dm = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', block)
                        if dm:
                            dock_date = dock_date or dm.group(1)
                        break
            # Case B: Quantity label appears, then a decimal on the next few lines and a unit token in vicinity
            if re.search(r'\bQuantity\b', line, re.IGNORECASE):
                area = lines[max(0, i-3):i+7]
                block = "\n".join(area)
                m = re.findall(r'\b(\d{1,3})\.00\b', block)
                if m:
                    vals = [int(x) for x in m if 1 <= int(x) <= 1000]
                    if vals:
                        quantity = min(vals)
                        dm = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', block)
                        if dm:
                            dock_date = dock_date or dm.group(1)
                        break
    
    # Strategy 4: Context-based extraction for part number lines (anywhere in text)
    if not quantity:
        for i, line in enumerate(lines):
            # Lines that contain likely part numbers
            if re.search(r'\b(?:\d{5,6}-\d+|[A-Z]{1,4}\d{2,6}(?:-\d+)?)\b', line, re.IGNORECASE):
                # This line likely contains item details including quantity
                
                # STRICT: Only look for quantities that end in .00
                decimal_matches = re.findall(r'(\d+)\.00\b', line)
                
                # Additional context: look for the smallest reasonable .00 value
                # (quantities are typically smaller than prices)
                valid_quantities = []
                for match in decimal_matches:
                    potential_qty = int(match)
                    if 1 <= potential_qty <= 100:  # Conservative range
                        valid_quantities.append(potential_qty)
                
                if valid_quantities:
                    # Pick the smallest valid quantity (most likely to be actual quantity)
                    quantity = min(valid_quantities)
                    
                    # Look for dock date in context
                    context_lines = lines[max(0, i-1):i+3]
                    for context_line in context_lines:
                        date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', context_line)
                        if date_match:
                            try:
                                from datetime import datetime
                                found_date = date_match.group(1)
                                date_obj = datetime.strptime(found_date, '%m/%d/%Y')
                                current_date = datetime.now()
                                # Future dates are more likely to be dock dates
                                if (date_obj - current_date).days >= -7:
                                    dock_date = found_date
                                    break
                            except:
                                pass
                    break
    
    # Strategy 5: Generic fallback with strict validation
    if not quantity:
        # Look for any .00 values but apply strict filtering
        # Split text into lines and check each line carefully
        for line in lines:
            # Skip header lines and look for data lines
            if re.search(r'\b(EA|LBS|PCS|EACH|PIECES?)\b', line, re.IGNORECASE):
                # This line contains units - likely a data line
                decimal_matches = re.findall(r'(\d+)\.00\b', line)
                
                for match in decimal_matches:
                    potential_qty = int(match)
                    # Very conservative range - typical PO quantities
                    if 1 <= potential_qty <= 50:
                        quantity = potential_qty
                        break
                
                if quantity:
                    break
    
    # Strategy 6: Separate date extraction if not found with quantity
    if not dock_date:
        # Find all dates and pick the most likely dock date
        date_matches = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', text)
        for date in date_matches:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date, '%m/%d/%Y')
                current_date = datetime.now()
                # Look for future dates (dock dates are typically future delivery dates)
                if (date_obj - current_date).days >= -7 and (date_obj - current_date).days <= 365:
                    dock_date = date
                    break
            except:
                pass
    
    return quantity, dock_date

def extract_payment_terms(text):
    """Extract payment terms and return (terms, non_standard_flag)"""
    standard_terms = "30 Days from Date of Invoice"
    
    # The text shows "30 Days from" on one line and "Date" and "of" "Invoice" on subsequent lines
    # Look for this pattern across multiple lines
    lines = text.split('\n')

    # First, anchor on the 'Payment terms' label and scan the following lines
    for i in range(len(lines)):
        line = lines[i]
        label_here = re.search(r'payment\s*terms\.?', line, re.IGNORECASE) is not None
        # Handle split label across two lines ("Payment" then "terms")
        if not label_here and i + 1 < len(lines):
            next_line = lines[i+1]
            label_here = ('payment' in line.lower() and 'terms' in next_line.lower())
        if label_here:
            block = ' '.join([l.strip() for l in lines[i:i+10] if l.strip()])
            if all(w in block.lower() for w in ['30', 'days', 'from', 'date', 'of', 'invoice']):
                return standard_terms, False
            # If immediate next line is like '30 Days', capture it but mark non-standard
            if i + 1 < len(lines):
                nxt = lines[i+1].strip()
                if re.fullmatch(r'30\s+Days', nxt, flags=re.IGNORECASE):
                    # Look ahead a few lines to see if the full standard phrase follows
                    lookahead = ' '.join([l.strip() for l in lines[i+1:i+8] if l.strip()]).lower()
                    if all(w in lookahead for w in ['30', 'days', 'from', 'date', 'of', 'invoice']):
                        return standard_terms, False
                    return '30 Days', True
    
    for i, line in enumerate(lines):
        if re.search(r'\b30\b', line) and 'days' in line.lower():
            # Check a wider window to accommodate word-wrapped tokens
            combined = ' '.join([l.strip() for l in lines[i:i+8] if l.strip()])
            if all(w in combined.lower() for w in ['30', 'days', 'from', 'date', 'of', 'invoice']):
                return standard_terms, False
    
    # Look for explicit payment terms section
    payment_pattern = r'Payment\s*terms[:\.]*\s*([^\n]+)'
    match = re.search(payment_pattern, text, re.IGNORECASE)
    if match:
        terms = match.group(1).strip()
        is_non_standard = "30 days from date of invoice" not in terms.lower()
        return terms, is_non_standard
    
    # Look for any mention of payment terms in broader context
    for i, line in enumerate(lines):
        if 'payment' in line.lower() and 'terms' in line.lower():
            # Try to build a contiguous phrase following the label
            combined = ' '.join([l.strip() for l in lines[i:i+6] if l.strip()])
            if '30 days' in combined.lower() and 'invoice' in combined.lower():
                return standard_terms, False
            line_clean = ' '.join(line.split())
            return line_clean, True  # Found but likely non-standard
    
    # Default case - if we found "30 Days from" pattern but incomplete, assume standard
    for line in lines:
        if '30' in line and 'days' in line.lower() and 'from' in line.lower():
            return standard_terms, False  # Assume standard terms
    
    return None, True  # No terms found, flag as non-standard

def extract_vendor_info(text):
    """Extract vendor information and return (vendor_name, non_tek_flag)"""
    lines = text.split('\n')
    vendor_name = None
    
    # Strategy 1: Look for known vendor patterns first
    known_vendors = [
        ("TEK ENTERPRISES", "TEK ENTERPRISES, INC."),
        # Add more vendors as needed
    ]
    
    for vendor_key, vendor_full in known_vendors:
        if vendor_key.upper() in text.upper():
            vendor_name = vendor_full
            break
    
    # Strategy 1b: Handle "Confirmed with <VENDOR>" phrasing
    if not vendor_name:
        m = re.search(r'confirmed\s+with\s+([A-Z0-9\s\.&,-]+)', text, re.IGNORECASE)
        if m:
            cand = m.group(1).strip()
            # Normalize common endings
            cand = re.sub(r'\s{2,}', ' ', cand)
            # Trim trailing address fragments
            cand = re.split(r'\s(Inc\.?|LLC|Corp\.?|Company)\b', cand, maxsplit=1, flags=re.IGNORECASE)[0].strip() + ('' if not re.search(r'\b(Inc\.?|LLC|Corp\.?|Company)\b', cand, re.IGNORECASE) else '')
            # If TEK appears, force canonical name
            if 'TEK' in cand.upper():
                vendor_name = 'TEK ENTERPRISES, INC.'

    # Strategy 2: Generic vendor extraction - look for vendor field patterns
    if not vendor_name:
        for i, line in enumerate(lines):
            line_upper = line.upper()
            # Look for vendor section indicators
            if any(indicator in line_upper for indicator in ['VENDOR ADDRESS', 'VENDOR', 'SUPPLIER', 'FROM:']):
                # Check next few lines for vendor name
                for j in range(i + 1, min(i + 4, len(lines))):
                    next_line = lines[j].strip()
                    # Skip obvious non-vendor entries
                    if (next_line and 
                        not any(word in next_line.lower() for word in ['address', 'phone', 'fax', 'number', 'email']) and
                        len(next_line) > 5 and 
                        not next_line.isdigit() and
                        not re.match(r'^\d+\s+(.*)', next_line)):  # Skip address numbers
                        
                        # Clean up the vendor name
                        vendor_name = next_line.strip()
                        # Remove common prefixes/suffixes
                        vendor_name = re.sub(r'^(To:|From:|Ship to:|Bill to:)\s*', '', vendor_name, flags=re.IGNORECASE)
                        break
                if vendor_name:
                    break
    
    # Strategy 3: Look for company patterns (all caps, INC, LLC, etc.)
    if not vendor_name:
        company_pattern = r'([A-Z\s&,\.]{10,50}(?:INC|LLC|CORP|LTD|COMPANY|ENTERPRISES)[A-Z\s,\.]*)'
        company_matches = re.findall(company_pattern, text)
        if company_matches:
            # Filter out obvious non-vendors
            for match in company_matches:
                match_clean = match.strip()
                if (len(match_clean) > 10 and 
                    not any(word in match_clean.lower() for word in ['purchase', 'order', 'agreement', 'terms'])):
                    vendor_name = match_clean
                    break
    
    # Strategy 4: Extract from "confirmed with" or similar patterns
    if not vendor_name:
        confirmed_pattern = r'confirmed\s+with\s+([A-Za-z\s&,\.]+)'
        confirmed_matches = re.findall(confirmed_pattern, text, re.IGNORECASE)
        if confirmed_matches:
            vendor_name = confirmed_matches[0].strip()
    
    if vendor_name:
        # Normalize TEK vendor name
        if 'TEK' in vendor_name.upper():
            vendor_name = 'TEK ENTERPRISES, INC.'
        # Determine if this is a non-TEK vendor
        is_non_tek = 'TEK ENTERPRISES' not in vendor_name.upper()
        return vendor_name, is_non_tek
    
    return None, True  # No vendor found, assume non-TEK
    
    return None, True  # No vendor found, flag as non-standard

def extract_buyer_name(text):
    """Extract buyer's name from the document"""
    lines = text.split('\n')

    # Try robust capture across newlines immediately following the label
    m_block = re.search(r'Buyer/phone\s+([A-Za-z][A-Za-z\-\.]+(?:\s+[A-Za-z][A-Za-z\-\.]+)+)\s*/', text, re.IGNORECASE)
    if m_block:
        return m_block.group(1).strip()
    
    # First, look for known buyers (configurable list)
    known_buyers = ["Nataly Hernandez", "Daniel Rodriguez"]
    for buyer in known_buyers:
        if buyer in text:
            return buyer

    # Generic approach: Look for buyer name pattern - appears after "Buyer/phone" field
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if 'buyer/phone' in line_lower:
            # Try inline: "Buyer/phone <Name> / <phone>"
            m = re.search(r'buyer/phone\s+([A-Za-z]+(?:\s+[A-Za-z\-]+)+)\s*/', line, re.IGNORECASE)
            if m:
                return m.group(1).strip()
            # Handle line breaks where name spans the next lines before the '/'
            name_parts = []
            for j in range(i + 1, min(i + 6, len(lines))):
                nxt = lines[j].strip()
                # Stop if we hit a slash (likely phone) or digits/email
                if '/' in nxt or re.search(r'\d|@', nxt):
                    break
                if nxt and re.fullmatch(r'[A-Za-z][A-Za-z\-\.]*', nxt):
                    name_parts.append(nxt)
                # Stop after capturing two parts (First Last)
                if len(name_parts) >= 2:
                    break
            if name_parts:
                return ' '.join(name_parts)

    # Alternative: Look for email patterns and extract name before @
    email_pattern = r'([A-Za-z\s]+)\s*[<(]?([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})[>)]?'
    email_matches = re.findall(email_pattern, text)
    for name, email in email_matches:
        name = name.strip()
        if len(name) > 5 and len(name.split()) >= 2:
            return name

    # Last resort: Look for patterns like "Name: [Person Name]"
    name_pattern = r'(?:Name|Contact|Buyer):\s*([A-Za-z\s\.\-]{5,30})'
    name_matches = re.findall(name_pattern, text, re.IGNORECASE)
    if name_matches:
        return name_matches[0].strip()

    # Look for any proper name pattern that appears multiple times (likely buyer name)
    name_pattern = r'([A-Z][a-z]+\s+[A-Z][a-z]+)'
    matches = re.findall(name_pattern, text)
    if matches:
        name_counts = {}
        for match in matches:
            match_lower = match.lower()
            if not any(word in match_lower for word in ['street', 'avenue', 'road', 'drive', 'california', 'hollywood', 'north', 'meggitt', 'enterprises', 'currency', 'buyer']):
                name_counts[match] = name_counts.get(match, 0) + 1
        if name_counts:
            return max(name_counts, key=name_counts.get)

    return None

def extract_dpas_ratings(text):
    """Extract DPAS ratings from the document"""
    dpas_ratings = []
    
    # Look for DPAS Rating: followed by the ratings
    lines = text.split('\n')
    
    for i, line in enumerate(lines):
        line_upper = line.upper()
        if 'DPAS' in line_upper and 'RATING' in line_upper:
            # Check the same line and next few lines for ratings
            search_text = line
            for j in range(i + 1, min(i + 3, len(lines))):
                search_text += ' ' + lines[j]
            
            # Look for DPAS patterns: DOA, DOC, DXA followed by numbers (4 characters total)
            ratings = re.findall(r'(D[OX][AC]\d+)', search_text.upper())
            dpas_ratings.extend(ratings)
            break
    
    # If not found in structured way, search more broadly
    if not dpas_ratings:
        # Look for DOA/DOC/DXA patterns anywhere in text (4 characters total)
        all_ratings = re.findall(r'(D[OX][AC]\d+)', text.upper())
        dpas_ratings.extend(all_ratings)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_ratings = []
    for rating in dpas_ratings:
        if rating not in seen:
            seen.add(rating)
            unique_ratings.append(rating)
    
    return unique_ratings if unique_ratings else None

def extract_quality_clauses(text):
    """Extract Quality Clauses (Q numbers with descriptions) from the document with enhanced classification"""
    quality_clauses = {}
    
    # Define Q clause business classification
    q_clause_classification = {
        # Auto-accept - Standard clauses that don't affect timesheet pricing
        "auto_accept": {
            "Q1": {"description": "QUALITY SYSTEMS REQUIREMENTS", "timesheet_impact": False, "action": "ACCEPT", "notes": "Standard quality compliance"},
            "Q5": {"description": "CERTIFICATION OF CONFORMANCE AND RECORD RETENTION", "timesheet_impact": False, "action": "ACCEPT", "notes": "Standard COC"},
            "Q26": {"description": "PACKING FOR SHIPMENT", "timesheet_impact": False, "action": "ACCEPT", "notes": "Standard packing"}
        },
        # Review required - May affect timesheet/pricing  
        "review_required": {
            "Q2": {"description": "SURVEILLANCE BY MEGGITT AND RIGHT OF ENTRY", "timesheet_impact": True, "action": "REVIEW", "notes": "Customer access required", "alert_level": "MEDIUM"},
            "Q9": {"description": "CORRECTIVE ACTION", "timesheet_impact": True, "action": "REVIEW", "notes": "CA documentation required", "alert_level": "MEDIUM"},
            "Q11": {"description": "SPECIAL PROCESS SOURCES REQUIRED", "timesheet_impact": True, "action": "REVIEW", "notes": "Verify certifications", "alert_level": "HIGH"},
            "Q13": {"description": "REPORT OF DISCREPANCY # Quality Notification (QN)", "timesheet_impact": True, "action": "REVIEW", "notes": "QN reporting required", "alert_level": "HIGH"},
            "Q14": {"description": "FOREIGN OBJECT DAMAGE (FOD)", "timesheet_impact": True, "action": "REVIEW", "notes": "FOD prevention measures", "alert_level": "MEDIUM"}
        },
        # Object to - Typically reject or negotiate
        "object_to": {
            "Q15": {"description": "ANTI-TERRORIST POLICY", "timesheet_impact": False, "action": "OBJECT", "notes": "Standard objection", "alert_level": "HIGH"},
            "Q32": {"description": "FLOWDOWN OF REQUIREMENTS [QUALITY AND ENVIRONMENTAL]", "timesheet_impact": False, "action": "OBJECT", "notes": "Flowdown too broad", "alert_level": "HIGH"},
            "Q33": {"description": "FAR and DOD FAR SUPPLEMENTAL FLOWDOWN PROVISIONS", "timesheet_impact": False, "action": "OBJECT", "notes": "FAR inappropriate for commercial", "alert_level": "CRITICAL"}
        }
    }
    
    # Flatten classification for description lookup
    known_clauses = {}
    for category, clauses in q_clause_classification.items():
        for q_number, info in clauses.items():
            known_clauses[q_number] = info["description"]
    
    # Extract Q numbers that actually appear in the text
    q_numbers_found = re.findall(r'(Q\d+)', text.upper())
    
    # Remove duplicates while preserving order
    seen = set()
    unique_q_numbers = []
    for q in q_numbers_found:
        if q not in seen:
            seen.add(q)
            unique_q_numbers.append(q)
    
    # Map found Q numbers to their descriptions
    for q_number in unique_q_numbers:
        if q_number in known_clauses:
            quality_clauses[q_number] = known_clauses[q_number]
        else:
            # For unknown Q numbers, try to extract description from context
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if q_number in line.upper():
                    # Try to get description from same line or following lines
                    description_parts = []
                    remaining = line.split(q_number, 1)
                    if len(remaining) > 1:
                        desc = remaining[1].strip()
                        if desc:
                            description_parts.append(desc)
                    
                    # Look at next couple lines for continuation
                    for j in range(i + 1, min(i + 3, len(lines))):
                        next_line = lines[j].strip()
                        if next_line and not re.match(r'^Q\d+', next_line.upper()) and len(next_line) < 60:
                            description_parts.append(next_line)
                        else:
                            break
                    
                    if description_parts:
                        quality_clauses[q_number] = ' '.join(description_parts).strip()
                    break
    
    # Classify the Q clauses for business processing
    q_clause_analysis = classify_q_clauses_for_business(quality_clauses, q_clause_classification)
    
    # Return the complete analysis structure expected by main()
    return {
        'raw_clauses': list(quality_clauses.keys()),
        'quality_clauses_dict': quality_clauses,
        'classified_clauses': q_clause_analysis,  # Pass the entire analysis result
        'summary': q_clause_analysis.get('summary', {}),
        'overall_status': q_clause_analysis.get('action_required', False)  # Use action_required as status
    }

def classify_q_clauses_for_business(quality_clauses_dict, classification):
    """Classify Q clauses into business categories for FileMaker integration"""
    if not quality_clauses_dict:
        return {
            "total_clauses": 0,
            "accept_clauses": [],
            "review_clauses": [], 
            "object_clauses": [],
            "unknown_clauses": [],
            "timesheet_impact": "NONE",
            "action_required": False,
            "summary": "No Q clauses found"
        }
    
    result = {
        "total_clauses": len(quality_clauses_dict),
        "accept_clauses": [],
        "review_clauses": [],
        "object_clauses": [], 
        "unknown_clauses": [],
        "timesheet_impact": "NONE",
        "action_required": False,
        "summary": ""
    }
    
    # Classify each Q clause
    for q_number, description in quality_clauses_dict.items():
        clause_data = {"number": q_number, "description": description}
        
        # Find in classification
        found = False
        for category, clauses in classification.items():
            if q_number in clauses:
                clause_info = clauses[q_number]
                clause_data.update(clause_info)
                
                if category == "auto_accept":
                    result["accept_clauses"].append(clause_data)
                elif category == "review_required":
                    result["review_clauses"].append(clause_data)
                    if clause_info.get("timesheet_impact"):
                        result["timesheet_impact"] = "MEDIUM" if result["timesheet_impact"] == "NONE" else "HIGH"
                elif category == "object_to":
                    result["object_clauses"].append(clause_data)
                    result["action_required"] = True
                
                found = True
                break
        
        if not found:
            result["unknown_clauses"].append(clause_data)
            result["action_required"] = True
            result["timesheet_impact"] = "HIGH"
    
    # Determine overall timesheet impact
    if result["review_clauses"] or result["unknown_clauses"]:
        if any(c.get("alert_level") == "HIGH" for c in result["review_clauses"]):
            result["timesheet_impact"] = "HIGH"
        elif any(c.get("alert_level") == "MEDIUM" for c in result["review_clauses"]):
            if result["timesheet_impact"] == "NONE":
                result["timesheet_impact"] = "MEDIUM"
    
    # Generate summary
    summary_parts = []
    if result["accept_clauses"]:
        summary_parts.append(f"{len(result['accept_clauses'])} auto-accept")
    if result["review_clauses"]:
        summary_parts.append(f"{len(result['review_clauses'])} need review")
    if result["object_clauses"]:
        summary_parts.append(f"{len(result['object_clauses'])} to object")
    if result["unknown_clauses"]:
        summary_parts.append(f"{len(result['unknown_clauses'])} unknown")
    
    result["summary"] = "; ".join(summary_parts) if summary_parts else "No clauses"
    
    return result

def extract_text_from_first_router_page(pdf_path):
    """Extract text from only the first page of the router PDF"""
    if not os.path.exists(pdf_path):
        return ""
    
    doc = fitz.open(pdf_path)
    if len(doc) == 0:
        doc.close()
        return ""
    
    # Get only first page (page 0)
    page = doc[0]
    
    # Try to extract text directly first
    page_text = page.get_text()
    
    # If no text found or minimal text, use OCR
    if len(page_text.strip()) < 50:
        try:
            # Use adaptive resolution for first router page
            matrices = [
                fitz.Matrix(3, 3),  # Standard high res
                fitz.Matrix(2, 2),  # Lower res for already high-quality scans
                fitz.Matrix(4, 4)   # Higher res for difficult scans
            ]
            
            best_text = ""
            best_length = 0
            
            for matrix in matrices:
                try:
                    pix = page.get_pixmap(matrix=matrix)
                    img_data = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_data)).convert('L')
                    
                    # Enhanced OCR configs for router pages
                    configs = [
                        r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz/-*().,: ',
                        r'--oem 3 --psm 4',
                        r'--oem 3 --psm 3',
                        r'--oem 1 --psm 6',
                        r'--oem 2 --psm 6'
                    ]
                    
                    for config in configs:
                        try:
                            ocr_text = pytesseract.image_to_string(img, config=config)
                            if len(ocr_text) > best_length:
                                best_text = ocr_text
                                best_length = len(ocr_text)
                        except:
                            continue
                    
                    # If we got decent results, don't try more matrices
                    if best_length > 100:
                        break
                        
                except Exception as e:
                    print(f"OCR matrix {matrix} failed for router page: {e}")
                    continue
            
            page_text = best_text if best_text else page_text
            
        except Exception as e:
            print(f"Error processing first router page: {e}")
    
    doc.close()
    return f"FIRST ROUTER PAGE:\n{page_text}\n\n"

def main():
    # Find the most recent PO folder (starts with 455)
    po_folders = [d for d in os.listdir('.') if os.path.isdir(d) and d.startswith('455')]
    if not po_folders:
        print("No PO folders found")
        return
    
    # Get the most recently modified PO folder
    po_folder = max(po_folders, key=lambda d: os.path.getmtime(d))
    po_number = po_folder
    po_file = os.path.join(po_folder, f"PO_{po_number}.pdf")
    router_file = os.path.join(po_folder, f"Router_{po_number}.pdf")
    json_file = os.path.join(po_folder, f"{po_number}_info.json")
    
    if not os.path.exists(po_file):
        print(f"PO file not found: {po_file}")
        return
    
    print("Extracting detailed text from ALL PO pages...")
    text = extract_text_from_pdf(po_file)
    
    # ENHANCEMENT: Also extract text from first page of router section
    if os.path.exists(router_file):
        print("Extracting text from first router page for additional information...")
        router_text = extract_text_from_first_router_page(router_file)
        text += router_text  # Combine PO text with first router page text
        print("Combined PO text with first router page text for comprehensive extraction")
    else:
        print("No router file found - processing only PO pages")
    
    print("\\nExtracting Production Order...")
    production_order = extract_production_order(text)
    print(f"Production Order: {production_order}")
    
    print("\\nExtracting Revision...")
    revision = extract_revision(text)
    print(f"Revision: {revision}")
    
    print("\\nExtracting Part Number...")
    part_number = extract_part_number(text, production_order)
    print(f"Part Number: {part_number}")
    
    print("\\nExtracting Quantity and Dock Date...")
    quantity, dock_date = extract_quantity_and_dock_date(text)
    print(f"Quantity: {quantity}")
    print(f"Dock Date: {dock_date}")
    
    print("\\nExtracting Payment Terms...")
    payment_terms, payment_terms_flag = extract_payment_terms(text)
    print(f"Payment Terms: {payment_terms}")
    print(f"Non-standard Payment Terms Flag: {payment_terms_flag}")
    
    print("\\nExtracting Vendor Information...")
    vendor_name, vendor_flag = extract_vendor_info(text)
    print(f"Vendor: {vendor_name}")
    print(f"Non-Tek Vendor Flag: {vendor_flag}")
    
    print("\\nExtracting Buyer Name...")
    buyer_name = extract_buyer_name(text)
    print(f"Buyer: {buyer_name}")
    
    print("\\nExtracting DPAS Ratings...")
    dpas_ratings = extract_dpas_ratings(text)
    print(f"DPAS Ratings: {dpas_ratings}")
    
    print("\\nExtracting Quality Clauses...")
    quality_clauses_analysis = extract_quality_clauses(text)
    quality_clauses = quality_clauses_analysis.get('quality_clauses_dict', {})
    print(f"Quality Clauses Found: {len(quality_clauses)} clauses")
    print(f"Classification Summary: {quality_clauses_analysis.get('summary', {})}")
    
    # Load existing JSON
    with open(json_file, 'r') as f:
        po_info = json.load(f)
    
    # Add new information including enhanced Q clause analysis
    po_info.update({
        "production_order": production_order,
        "revision": revision,
        "part_number": part_number,
        "quantity": quantity,
        "dock_date": dock_date,
        "payment_terms": payment_terms,
        "payment_terms_non_standard_flag": payment_terms_flag,
        "vendor_name": vendor_name,
        "vendor_non_tek_flag": vendor_flag,
        "buyer_name": buyer_name,
        "dpas_ratings": dpas_ratings,
        "quality_clauses": quality_clauses,
        "quality_clauses_analysis": quality_clauses_analysis
    })
    
    # Save updated JSON
    with open(json_file, 'w') as f:
        json.dump(po_info, f, indent=2)
    
    print(f"\\nUpdated JSON file: {json_file}")
    print("\\nExtracted information:")
    print(json.dumps(po_info, indent=2))
    
    # Also save extracted text for debugging (now includes router first page if available)
    debug_file = os.path.join(po_folder, "extracted_text_comprehensive.txt")
    with open(debug_file, 'w', encoding='utf-8') as f:
        f.write("=== COMPREHENSIVE TEXT EXTRACTION ===\n")
        f.write("This includes ALL PO pages AND first router page (if available)\n")
        f.write("=" * 60 + "\n\n")
        f.write(text)
    print(f"\\nSaved comprehensive extracted text to: {debug_file}")
    
    # Move original searchable PDF to PO folder for complete organization
    original_pdf = po_info.get("source_file", "final_searchable_output.pdf")
    if os.path.exists(original_pdf):
        import shutil
        destination = os.path.join(po_folder, os.path.basename(original_pdf))
        try:
            shutil.move(original_pdf, destination)
            print(f"\\nMoved original PDF to: {destination}")
            # Update JSON to reflect new location
            po_info["source_file"] = os.path.basename(original_pdf)
            with open(json_file, 'w') as f:
                json.dump(po_info, f, indent=2)
        except Exception as e:
            print(f"\\nWarning: Could not move original PDF: {e}")
    else:
        print(f"\\nWarning: Original PDF not found: {original_pdf}")
    
    print(f"\\nComplete processing finished for PO {po_number}")
    print(f"All files organized in folder: {po_folder}")

if __name__ == "__main__":
    main()